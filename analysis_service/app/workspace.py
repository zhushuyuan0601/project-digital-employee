from __future__ import annotations

import csv
import base64
import html
import io
import json
import os
import re
import shutil
import sqlite3
import time
import xml.dom.minidom
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from fastapi.responses import FileResponse

from .storage import WORKSPACE_ROOT, workspace_dir

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None

try:
    import yaml  # type: ignore
except Exception:
    yaml = None


TEXT_EXTENSIONS = {".txt", ".md", ".markdown", ".log", ".json", ".py", ".sql", ".xml", ".yaml", ".yml"}
TABLE_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}
SQLITE_EXTENSIONS = {".sqlite", ".db"}
BLOCKED_UPLOAD_EXTENSIONS = {".py"}
MAX_UPLOAD_FILES = int(os.getenv("ANALYSIS_MAX_UPLOAD_FILES", "10"))
MAX_UPLOAD_FILE_SIZE = int(os.getenv("ANALYSIS_MAX_UPLOAD_FILE_SIZE", str(50 * 1024 * 1024)))
MAX_UPLOAD_TOTAL_SIZE = int(os.getenv("ANALYSIS_MAX_UPLOAD_TOTAL_SIZE", str(200 * 1024 * 1024)))
PROFILE_SAMPLE_SIZE = 1000
PROFILE_FIELD_LIMIT = 40
PROFILE_SAMPLE_VALUE_LIMIT = 5
TEXT_PREVIEW_LIMIT = int(os.getenv("ANALYSIS_TEXT_PREVIEW_LIMIT", str(200 * 1024)))
DELIMITED_TEXT_MIN_TABS = 2
CONTEXT_CELL_LIMIT = 240

_PROFILE_CACHE: dict[str, dict[str, Any]] = {}


def is_within_workspace(path: Path, session_id: str) -> bool:
    root = workspace_dir(session_id).resolve()
    try:
        resolved = path.resolve()
    except Exception:
        return False
    return resolved == root or root in resolved.parents


def resolve_session_path(session_id: str, rel_path: str) -> Path:
    root = workspace_dir(session_id)
    candidate = (root / rel_path).resolve()
    if not is_within_workspace(candidate, session_id):
        raise HTTPException(status_code=400, detail="Invalid workspace path")
    return candidate


def classify_file(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in TABLE_EXTENSIONS or ext in SQLITE_EXTENSIONS or is_delimited_text_file(path):
        return "table"
    if ext in IMAGE_EXTENSIONS:
        return "image"
    return "other"


def classify_artifact(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in IMAGE_EXTENSIONS:
        return "chart"
    if ext in TABLE_EXTENSIONS or ext in SQLITE_EXTENSIONS:
        return "table"
    if ext in {".md", ".markdown", ".pdf", ".doc", ".docx", ".html"} or any(
        marker in path.name.lower() for marker in ("report", "summary", "分析报告", "报告")
    ):
        return "report"
    return "other"


def is_delimited_text_file(path: Path) -> bool:
    if path.suffix.lower() not in {".txt", ".log"}:
        return False
    try:
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            checked = 0
            tabbed = 0
            for line in handle:
                stripped = line.strip()
                if not stripped:
                    continue
                checked += 1
                if stripped.count("\t") >= DELIMITED_TEXT_MIN_TABS:
                    tabbed += 1
                if checked >= 8:
                    break
        return checked > 0 and tabbed / checked >= 0.6
    except Exception:
        return False


def _delimited_text_headers(width: int) -> list[str]:
    common = ["tag", "call_id", "caller_utterance", "ai_utterance", "dialogue", "date"]
    if width <= len(common):
        return common[:width]
    return common + [f"field_{index + 1}" for index in range(len(common), width)]


def _compact_cell(value: Any, limit: int = CONTEXT_CELL_LIMIT) -> str:
    text = str(value if value is not None else "").replace("\n", " ").strip()
    return text[:limit] + ("..." if len(text) > limit else "")


def _compact_rows(rows: list[list[Any]], limit: int = CONTEXT_CELL_LIMIT) -> list[list[str]]:
    return [[_compact_cell(cell, limit) for cell in row] for row in rows]


def build_download_url(session_id: str, rel_path: str) -> str:
    return f"/workspace/download?session_id={session_id}&path={rel_path}"


def build_preview_url(session_id: str, rel_path: str) -> str:
    return f"/workspace/preview?session_id={session_id}&path={rel_path}"


def list_workspace_files(session_id: str) -> list[dict[str, Any]]:
    root = workspace_dir(session_id)
    files: list[dict[str, Any]] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.name.startswith("."):
            continue
        rel = path.relative_to(root).as_posix()
        files.append({
            "name": path.name,
            "path": rel,
            "size": path.stat().st_size,
            "category": classify_file(path),
            "artifact_type": classify_artifact(path) if rel.startswith("generated/") else "source",
            "is_generated": rel.startswith("generated/"),
            "download_url": build_download_url(session_id, rel),
            "preview_url": build_preview_url(session_id, rel),
        })
    return files


def _tree_node(path: Path, root: Path, session_id: str) -> dict[str, Any]:
    rel = "." if path == root else path.relative_to(root).as_posix()
    if path.is_dir():
        return {
            "name": path.name if path != root else session_id,
            "path": "" if rel == "." else rel,
            "type": "directory",
            "children": [
                _tree_node(child, root, session_id)
                for child in sorted(path.iterdir(), key=lambda item: (not item.is_dir(), item.name.lower()))
                if not child.name.startswith(".")
            ],
        }
    return {
        "name": path.name,
        "path": rel,
        "type": "file",
        "category": classify_file(path),
        "download_url": build_download_url(session_id, rel),
        "preview_url": build_preview_url(session_id, rel),
        "is_generated": rel.startswith("generated/"),
    }


def build_tree(session_id: str) -> dict[str, Any]:
    root = workspace_dir(session_id)
    return _tree_node(root, root, session_id)


async def upload_files(session_id: str, files: list[UploadFile], directory: str = "") -> list[dict[str, Any]]:
    if len(files) > MAX_UPLOAD_FILES:
        raise HTTPException(status_code=413, detail=f"Too many files. Maximum is {MAX_UPLOAD_FILES}")
    target_dir = resolve_session_path(session_id, directory) if directory else workspace_dir(session_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    uploaded: list[dict[str, Any]] = []
    total_size = 0
    for upload in files:
        filename = Path(upload.filename or "upload.bin").name
        if Path(filename).suffix.lower() in BLOCKED_UPLOAD_EXTENSIONS:
            raise HTTPException(status_code=400, detail=f"Blocked upload type: {filename}")
        target = target_dir / filename
        counter = 1
        while target.exists():
            target = target_dir / f"{target.stem}_{counter}{target.suffix}"
            counter += 1
        size = 0
        with target.open("wb") as handle:
            while True:
                chunk = await upload.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                total_size += len(chunk)
                if size > MAX_UPLOAD_FILE_SIZE:
                    target.unlink(missing_ok=True)
                    raise HTTPException(status_code=413, detail=f"File too large: {filename}")
                if total_size > MAX_UPLOAD_TOTAL_SIZE:
                    target.unlink(missing_ok=True)
                    raise HTTPException(status_code=413, detail="Upload total size exceeded")
                handle.write(chunk)
        uploaded.append({
            "name": target.name,
            "path": target.relative_to(workspace_dir(session_id)).as_posix(),
            "size": target.stat().st_size,
        })
    return uploaded


def delete_path(session_id: str, rel_path: str) -> None:
    target = resolve_session_path(session_id, rel_path)
    if not target.exists():
        raise HTTPException(status_code=404, detail="Workspace file not found")
    if target.is_dir():
        shutil.rmtree(target)
    else:
        target.unlink()


def clear_workspace(session_id: str) -> None:
    root = workspace_dir(session_id)
    for child in root.iterdir():
        if child.name.startswith("."):
            continue
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()


def download_file(session_id: str, rel_path: str) -> FileResponse:
    target = resolve_session_path(session_id, rel_path)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="Workspace file not found")
    return FileResponse(target, filename=target.name)


def _preview_text(path: Path) -> dict[str, Any]:
    truncated = path.stat().st_size > TEXT_PREVIEW_LIMIT
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        content = handle.read(TEXT_PREVIEW_LIMIT + 1)
    if len(content) > TEXT_PREVIEW_LIMIT:
        content = content[:TEXT_PREVIEW_LIMIT]
        truncated = True
    return {"kind": "text", "content": content, "truncated": truncated, "size": path.stat().st_size}


def _html_page(title: str, body: str, *, subtitle: str = "") -> str:
    safe_title = html.escape(title)
    safe_subtitle = html.escape(subtitle)
    subtitle_markup = f"<p class=\"subtitle\">{safe_subtitle}</p>" if safe_subtitle else ""
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{safe_title}</title>
  <style>
    :root {{
      color-scheme: light;
      --ink: #172033;
      --muted: #667085;
      --line: #d9e2ec;
      --surface: #ffffff;
      --accent: #2563eb;
      --soft: #f6f8fb;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: #eef2f7;
      color: var(--ink);
      font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      line-height: 1.68;
    }}
    main {{
      width: min(1120px, calc(100vw - 32px));
      margin: 32px auto;
      padding: 32px;
      border: 1px solid var(--line);
      border-radius: 12px;
      background: var(--surface);
      box-shadow: 0 24px 80px rgba(15, 23, 42, 0.08);
    }}
    header {{ margin-bottom: 24px; border-bottom: 1px solid var(--line); padding-bottom: 18px; }}
    h1 {{ margin: 0; font-size: clamp(24px, 4vw, 42px); line-height: 1.12; letter-spacing: 0; }}
    h2 {{ margin-top: 28px; font-size: 22px; line-height: 1.28; }}
    h3 {{ margin-top: 22px; font-size: 18px; }}
    .subtitle {{ margin: 10px 0 0; color: var(--muted); font-size: 14px; }}
    pre {{
      overflow: auto;
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #0f172a;
      color: #e5e7eb;
      line-height: 1.55;
    }}
    code {{ font-family: "SFMono-Regular", Consolas, "Liberation Mono", monospace; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ padding: 10px 12px; border-bottom: 1px solid var(--line); text-align: left; vertical-align: top; }}
    th {{ position: sticky; top: 0; background: var(--soft); color: #344054; font-weight: 700; }}
    .table-wrap {{ overflow: auto; max-height: 70vh; border: 1px solid var(--line); border-radius: 10px; }}
    .metric-row {{ display: flex; flex-wrap: wrap; gap: 10px; margin: 0 0 18px; }}
    .metric {{ padding: 10px 12px; border: 1px solid var(--line); border-radius: 10px; background: var(--soft); }}
    .metric strong {{ display: block; font-size: 20px; line-height: 1.1; }}
    img {{ display: block; max-width: 100%; height: auto; border-radius: 10px; border: 1px solid var(--line); }}
    a {{ color: var(--accent); }}
    @media (max-width: 720px) {{
      main {{ width: min(100vw - 16px, 1120px); margin: 8px auto; padding: 18px; border-radius: 10px; }}
      th, td {{ padding: 8px; }}
    }}
  </style>
</head>
<body>
  <main>
    <header>
      <h1>{safe_title}</h1>
      {subtitle_markup}
    </header>
    {body}
  </main>
</body>
</html>"""


def _markdown_to_html(content: str) -> str:
    lines = content.splitlines()
    html_lines: list[str] = []
    in_list = False
    in_code = False
    code_lines: list[str] = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("```"):
            if in_code:
                html_lines.append(f"<pre><code>{html.escape(chr(10).join(code_lines))}</code></pre>")
                code_lines = []
                in_code = False
            else:
                in_code = True
            continue
        if in_code:
            code_lines.append(line)
            continue
        if not stripped:
            if in_list:
                html_lines.append("</ul>")
                in_list = False
            continue
        heading = re.match(r"^(#{1,3})\s+(.+)$", stripped)
        if heading:
            if in_list:
                html_lines.append("</ul>")
                in_list = False
            level = len(heading.group(1))
            html_lines.append(f"<h{level}>{html.escape(heading.group(2))}</h{level}>")
            continue
        if stripped.startswith(("- ", "* ")):
            if not in_list:
                html_lines.append("<ul>")
                in_list = True
            html_lines.append(f"<li>{html.escape(stripped[2:])}</li>")
            continue
        if in_list:
            html_lines.append("</ul>")
            in_list = False
        html_lines.append(f"<p>{html.escape(stripped)}</p>")
    if in_code:
        html_lines.append(f"<pre><code>{html.escape(chr(10).join(code_lines))}</code></pre>")
    if in_list:
        html_lines.append("</ul>")
    return "\n".join(html_lines) or "<p>该文件暂无内容。</p>"


def _preview_html_text(path: Path) -> dict[str, Any]:
    preview = _preview_text(path)
    content = str(preview.get("content") or "")
    ext = path.suffix.lower()
    if ext in {".md", ".markdown"}:
        body = _markdown_to_html(content)
    elif ext in {".json"}:
        try:
            formatted = json.dumps(json.loads(content), ensure_ascii=False, indent=2)
        except Exception:
            formatted = content
        body = f"<pre><code>{html.escape(formatted)}</code></pre>"
    else:
        body = f"<pre><code>{html.escape(content)}</code></pre>"
    notice = ""
    if preview.get("truncated"):
        notice = "<p class=\"subtitle\">文件较大，当前仅展示前 200KB 内容。请下载查看完整文件。</p>"
    return {"kind": "html", "content": _html_page(path.name, notice + body, subtitle="HTML preview generated from workspace asset"), "truncated": preview.get("truncated", False), "size": preview.get("size")}


def _preview_html_table(payload: dict[str, Any], title: str) -> dict[str, Any]:
    columns = [str(column) for column in payload.get("columns") or []]
    rows = payload.get("rows") or []
    header = "".join(f"<th>{html.escape(column)}</th>" for column in columns)
    body_rows = []
    for row in rows:
        cells = row if isinstance(row, list) else list(row or [])
        body_rows.append("<tr>" + "".join(f"<td>{html.escape(str(cell))}</td>" for cell in cells) + "</tr>")
    metrics = (
        "<div class=\"metric-row\">"
        f"<div class=\"metric\"><strong>{html.escape(str(payload.get('row_count', len(rows))))}</strong><span>总行数</span></div>"
        f"<div class=\"metric\"><strong>{html.escape(str(len(columns)))}</strong><span>字段数</span></div>"
        f"<div class=\"metric\"><strong>{html.escape(str(len(rows)))}</strong><span>当前预览行</span></div>"
        "</div>"
    )
    table = f"<div class=\"table-wrap\"><table><thead><tr>{header}</tr></thead><tbody>{''.join(body_rows)}</tbody></table></div>"
    return {"kind": "html", "content": _html_page(title, metrics + table, subtitle="HTML table preview generated from workspace asset")}


def _preview_html_image(session_id: str, rel_path: str, path: Path) -> dict[str, Any]:
    mime = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".svg": "image/svg+xml",
    }.get(path.suffix.lower(), "application/octet-stream")
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    body = f"<figure><img src=\"data:{mime};base64,{encoded}\" alt=\"{html.escape(path.name)}\" /></figure>"
    return {"kind": "html", "content": _html_page(path.name, body, subtitle=f"Workspace path: {html.escape(rel_path)}")}


def _preview_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {"kind": "text", "content": json.dumps(payload, ensure_ascii=False, indent=2)}


def _preview_yaml(path: Path) -> dict[str, Any]:
    if yaml is None:
        return _preview_text(path)
    payload = yaml.safe_load(path.read_text(encoding="utf-8"))
    return {"kind": "text", "content": json.dumps(payload, ensure_ascii=False, indent=2)}


def _preview_xml(path: Path) -> dict[str, Any]:
    xml_str = path.read_text(encoding="utf-8", errors="replace")
    try:
        pretty = xml.dom.minidom.parseString(xml_str.encode("utf-8")).toprettyxml(indent="  ")
    except Exception:
        pretty = xml_str
    return {"kind": "text", "content": pretty}


def _table_payload(headers: list[str], rows: list[list[Any]], total_rows: int) -> dict[str, Any]:
    return {
        "kind": "table",
        "columns": headers,
        "rows": rows,
        "row_count": total_rows,
        "truncated": total_rows > len(rows),
    }


def _table_delimiter(path: Path) -> str:
    return "\t" if path.suffix.lower() == ".tsv" else ","


def _preview_csv(path: Path, page: int, page_size: int) -> dict[str, Any]:
    delimiter = _table_delimiter(path)
    start = max(page - 1, 0) * page_size
    end = start + page_size
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        headers = next(reader, [])
        records: list[list[Any]] = []
        total_rows = 0
        for total_rows, row in enumerate(reader, start=1):
            row_index = total_rows - 1
            if start <= row_index < end:
                records.append(row)
    return _table_payload(headers, records, total_rows)


def _preview_delimited_text(path: Path, page: int, page_size: int) -> dict[str, Any]:
    start = max(page - 1, 0) * page_size
    end = start + page_size
    records: list[list[Any]] = []
    total_rows = 0
    max_width = 0
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for total_rows, row in enumerate(reader, start=1):
            max_width = max(max_width, len(row))
            row_index = total_rows - 1
            if start <= row_index < end:
                records.append(row)
    return _table_payload(_delimited_text_headers(max_width), records, total_rows)


def _preview_excel(path: Path, page: int, page_size: int, sheet_name: str = "") -> dict[str, Any]:
    if openpyxl is None:
        return {
            "kind": "unsupported",
            "content": "openpyxl is not installed; Excel preview is unavailable.",
        }
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        target_name = sheet_name if sheet_name and sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[target_name]
        iterator = sheet.iter_rows(values_only=True)
        first = next(iterator, None)
        headers = [str(item) if item is not None else "" for item in first] if first else []
        start = max(page - 1, 0) * page_size
        end = start + page_size
        records: list[list[Any]] = []
        total_rows = 0
        for total_rows, row in enumerate(iterator, start=1):
            row_index = total_rows - 1
            if start <= row_index < end:
                records.append(["" if value is None else str(value) for value in row])
        payload = _table_payload(headers, records, total_rows)
        payload["sheet_name"] = target_name
        payload["sheet_names"] = workbook.sheetnames
        return payload
    finally:
        workbook.close()


def _preview_sqlite(path: Path, page: int, page_size: int, table_name: str = "") -> dict[str, Any]:
    connection = sqlite3.connect(path)
    try:
        tables = [
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' ORDER BY name"
            ).fetchall()
        ]
        if not tables:
            return {"kind": "database", "view": "tables", "tables": []}
        target = table_name if table_name and table_name in tables else tables[0]
        cursor = connection.execute(f'SELECT * FROM "{target}" LIMIT ? OFFSET ?', (page_size, (page - 1) * page_size))
        rows = cursor.fetchall()
        headers = [desc[0] for desc in cursor.description or []]
        total = connection.execute(f'SELECT COUNT(*) FROM "{target}"').fetchone()[0]
        return {
            "kind": "database",
            "view": "table",
            "table_name": target,
            "tables": tables,
            "columns": headers,
            "rows": rows,
            "row_count": total,
            "truncated": total > len(rows),
        }
    finally:
        connection.close()


def preview_file(session_id: str, rel_path: str, page: int = 1, page_size: int = 50, table_name: str = "", sheet_name: str = "") -> dict[str, Any]:
    path = resolve_session_path(session_id, rel_path)
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Workspace file not found")
    ext = path.suffix.lower()
    if rel_path.startswith("generated/") or classify_artifact(path) != "other":
        if is_delimited_text_file(path):
            return _preview_html_table(_preview_delimited_text(path, page, page_size), path.name)
        if ext in TEXT_EXTENSIONS:
            return _preview_html_text(path)
        if ext in {".csv", ".tsv"}:
            return _preview_html_table(_preview_csv(path, page, page_size), path.name)
        if ext in {".xlsx", ".xls"}:
            return _preview_html_table(_preview_excel(path, page, page_size, sheet_name=sheet_name), path.name)
        if ext in SQLITE_EXTENSIONS:
            return _preview_html_table(_preview_sqlite(path, page, page_size, table_name=table_name), path.name)
        if ext in IMAGE_EXTENSIONS:
            return _preview_html_image(session_id, rel_path, path)
    if ext in {".json"}:
        return _preview_json(path)
    if ext in {".yaml", ".yml"}:
        return _preview_yaml(path)
    if ext in {".xml"}:
        return _preview_xml(path)
    if ext in TEXT_EXTENSIONS:
        if is_delimited_text_file(path):
            return _preview_delimited_text(path, page, page_size)
        return _preview_text(path)
    if ext in {".csv", ".tsv"}:
        return _preview_csv(path, page, page_size)
    if ext in {".xlsx", ".xls"}:
        return _preview_excel(path, page, page_size, sheet_name=sheet_name)
    if ext in SQLITE_EXTENSIONS:
        return _preview_sqlite(path, page, page_size, table_name=table_name)
    if ext in IMAGE_EXTENSIONS:
        return {"kind": "image", "url": build_download_url(session_id, rel_path)}
    return {"kind": "unsupported", "content": f"Preview is not available for {path.name}"}


def _is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _infer_field_type(values: list[Any]) -> str:
    sample = [value for value in values if not _is_empty(value)][:200]
    if not sample:
        return "空值"
    numeric_count = 0
    date_count = 0
    for value in sample:
        text = str(value).strip().replace(",", "")
        try:
            float(text)
            numeric_count += 1
        except Exception:
            pass
        if any(mark in str(value) for mark in ("-", "/", "年", "月", "日")):
            try:
                normalized = str(value).replace("年", "-").replace("月", "-").replace("日", "")
                datetime_value = __import__("datetime").datetime
                datetime_value.fromisoformat(normalized.split(" ")[0])
                date_count += 1
            except Exception:
                pass
    if numeric_count / len(sample) >= 0.85:
        return "数值"
    if date_count / len(sample) >= 0.65:
        return "日期"
    unique_count = len({str(value) for value in sample})
    if unique_count <= max(8, len(sample) * 0.45):
        return "分类"
    return "文本"


def _numeric_stats(values: list[Any]) -> dict[str, Any] | None:
    numbers: list[float] = []
    for value in values:
        if _is_empty(value):
            continue
        try:
            numbers.append(float(str(value).replace(",", "")))
        except Exception:
            continue
    if not numbers:
        return None
    return {
        "min": min(numbers),
        "max": max(numbers),
        "mean": sum(numbers) / len(numbers),
    }


def _profile_table(
    headers: list[str],
    records: list[list[Any]],
    *,
    row_count: int | None = None,
    sampled_rows: int | None = None,
) -> dict[str, Any]:
    normalized_headers = [str(header) if str(header).strip() else f"字段 {index + 1}" for index, header in enumerate(headers)]
    sample_row_count = len(records)
    total_row_count = row_count if row_count is not None else sample_row_count
    column_count = len(normalized_headers)
    total_cells = max(sample_row_count * column_count, 1)
    missing_total = 0
    fields: list[dict[str, Any]] = []
    for index, name in enumerate(normalized_headers):
        values = [row[index] if index < len(row) else "" for row in records]
        missing_count = sum(1 for value in values if _is_empty(value))
        missing_total += missing_count
        non_empty = [str(value) for value in values if not _is_empty(value)]
        stats = _numeric_stats(values)
        field: dict[str, Any] = {
            "name": name,
            "type": _infer_field_type(values),
            "missing_count": missing_count,
            "missing_rate": missing_count / max(sample_row_count, 1),
            "unique_count": len(set(non_empty)),
            "samples": non_empty[:PROFILE_SAMPLE_VALUE_LIMIT],
        }
        if stats:
            field["stats"] = stats
        fields.append(field)
    return {
        "kind": "table_profile",
        "row_count": total_row_count,
        "column_count": column_count,
        "missing_count": missing_total,
        "missing_rate": missing_total / total_cells,
        "sampled_rows": sampled_rows if sampled_rows is not None else sample_row_count,
        "sampling": {
            "mode": "head",
            "sample_size": sampled_rows if sampled_rows is not None else sample_row_count,
            "is_sampled": total_row_count > sample_row_count,
        },
        "fields": fields,
    }


def _profile_csv(path: Path) -> dict[str, Any]:
    delimiter = _table_delimiter(path)
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        headers = next(reader, [])
        records: list[list[Any]] = []
        total_rows = 0
        for total_rows, row in enumerate(reader, start=1):
            if len(records) < PROFILE_SAMPLE_SIZE:
                records.append(row)
    return _profile_table(headers, records, row_count=total_rows, sampled_rows=len(records))


def _profile_delimited_text(path: Path) -> dict[str, Any]:
    records: list[list[Any]] = []
    total_rows = 0
    max_width = 0
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for total_rows, row in enumerate(reader, start=1):
            max_width = max(max_width, len(row))
            if len(records) < PROFILE_SAMPLE_SIZE:
                records.append(row)
    return _profile_table(_delimited_text_headers(max_width), records, row_count=total_rows, sampled_rows=len(records))


def _profile_excel(path: Path, sheet_name: str = "") -> dict[str, Any]:
    if openpyxl is None:
        return {"kind": "unsupported", "content": "openpyxl is not installed; Excel profile is unavailable."}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        target_name = sheet_name if sheet_name and sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[target_name]
        iterator = sheet.iter_rows(values_only=True)
        first = next(iterator, None)
        headers = [str(item) if item is not None else "" for item in first] if first else []
        records: list[list[Any]] = []
        total_rows = 0
        for total_rows, row in enumerate(iterator, start=1):
            if len(records) < PROFILE_SAMPLE_SIZE:
                records.append([cell if cell is not None else "" for cell in row])
        payload = _profile_table(headers, records, row_count=total_rows, sampled_rows=len(records))
        payload["sheet_name"] = target_name
        payload["sheet_names"] = workbook.sheetnames
        return payload
    finally:
        workbook.close()


def _profile_sqlite(path: Path, table_name: str = "") -> dict[str, Any]:
    connection = sqlite3.connect(path)
    try:
        tables = [
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' ORDER BY name"
            ).fetchall()
        ]
        if not tables:
            return {"kind": "database_profile", "tables": [], "row_count": 0, "column_count": 0, "fields": []}
        target = table_name if table_name and table_name in tables else tables[0]
        cursor = connection.execute(f'SELECT * FROM "{target}" LIMIT ?', (PROFILE_SAMPLE_SIZE,))
        headers = [desc[0] for desc in cursor.description or []]
        rows = cursor.fetchall()
        total = connection.execute(f'SELECT COUNT(*) FROM "{target}"').fetchone()[0]
        payload = _profile_table(headers, rows, row_count=total, sampled_rows=len(rows))
        payload["kind"] = "database_profile"
        payload["table_name"] = target
        payload["tables"] = tables
        return payload
    finally:
        connection.close()


def profile_file(session_id: str, rel_path: str, table_name: str = "", sheet_name: str = "") -> dict[str, Any]:
    path = resolve_session_path(session_id, rel_path)
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Workspace file not found")
    stat = path.stat()
    cache_key = f"{session_id}:{rel_path}:{stat.st_mtime_ns}:{stat.st_size}:{table_name}:{sheet_name}"
    cached = _PROFILE_CACHE.get(cache_key)
    if cached:
        return cached
    ext = path.suffix.lower()
    if is_delimited_text_file(path):
        payload = _profile_delimited_text(path)
    elif ext in {".csv", ".tsv"}:
        payload = _profile_csv(path)
    elif ext in {".xlsx", ".xls"}:
        payload = _profile_excel(path, sheet_name=sheet_name)
    elif ext in SQLITE_EXTENSIONS:
        payload = _profile_sqlite(path, table_name=table_name)
    else:
        payload = {
            "kind": "file_profile",
            "name": path.name,
            "size": stat.st_size,
            "category": classify_file(path),
        }
    payload.update({
        "source_path": rel_path,
        "profile_generated_at": int(time.time()),
    })
    _PROFILE_CACHE[cache_key] = payload
    if len(_PROFILE_CACHE) > 128:
        _PROFILE_CACHE.pop(next(iter(_PROFILE_CACHE)))
    return payload


def _peek_csv(path: Path, max_rows: int = 5) -> dict[str, Any] | None:
    try:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
            reader = csv.reader(handle, delimiter=delimiter)
            headers = next(reader, None)
            if not headers:
                return None
            rows: list[list[str]] = []
            for _, row in zip(range(max_rows), reader):
                rows.append(row)
        return {"columns": headers, "sample_rows": _compact_rows(rows), "total_rows": None}
    except Exception:
        return None


def _peek_excel(path: Path, max_rows: int = 5) -> dict[str, Any] | None:
    if openpyxl is None:
        return None
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        sheet = workbook[sheet_names[0]]
        collected: list[list[str]] = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            collected.append([str(cell) if cell is not None else "" for cell in row])
            if i >= max_rows:
                break
        workbook.close()
        if not collected:
            return None
        return {
            "columns": collected[0],
            "sample_rows": _compact_rows(collected[1:]),
            "sheet_names": sheet_names,
            "total_rows": None,
        }
    except Exception:
        return None


def _peek_sqlite(path: Path, max_rows: int = 5) -> dict[str, Any] | None:
    try:
        connection = sqlite3.connect(path)
        tables = [
            row[0] for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' ORDER BY name"
            ).fetchall()
        ]
        if not tables:
            connection.close()
            return None
        target = tables[0]
        cursor = connection.execute(f'SELECT * FROM "{target}" LIMIT ?', (max_rows,))
        headers = [desc[0] for desc in cursor.description or []]
        rows = [[str(cell) if cell is not None else "" for cell in row] for row in cursor.fetchall()]
        total = connection.execute(f'SELECT COUNT(*) FROM "{target}"').fetchone()[0]
        connection.close()
        return {
            "columns": headers,
            "sample_rows": _compact_rows(rows),
            "tables": tables,
            "total_rows": total,
        }
    except Exception:
        return None


def _compact_profile(profile: dict[str, Any], field_limit: int = PROFILE_FIELD_LIMIT) -> dict[str, Any]:
    fields = profile.get("fields") if isinstance(profile.get("fields"), list) else []
    compact_fields: list[dict[str, Any]] = []
    for field in fields[:field_limit]:
        compact_field = {
            "name": field.get("name"),
            "type": field.get("type"),
            "missing_rate": field.get("missing_rate"),
            "unique_count": field.get("unique_count"),
            "samples": [_compact_cell(sample) for sample in field.get("samples", [])[:PROFILE_SAMPLE_VALUE_LIMIT]],
        }
        if field.get("stats"):
            compact_field["stats"] = field.get("stats")
        compact_fields.append(compact_field)
    payload = {
        "kind": profile.get("kind"),
        "row_count": profile.get("row_count"),
        "column_count": profile.get("column_count"),
        "missing_rate": profile.get("missing_rate"),
        "sampled_rows": profile.get("sampled_rows"),
        "sheet_name": profile.get("sheet_name"),
        "sheet_names": profile.get("sheet_names"),
        "table_name": profile.get("table_name"),
        "tables": profile.get("tables"),
        "fields": compact_fields,
    }
    return {key: value for key, value in payload.items() if value not in (None, "", [])}


def _field_names(profile: dict[str, Any]) -> set[str]:
    fields = profile.get("fields") if isinstance(profile.get("fields"), list) else []
    return {str(field.get("name") or "").strip() for field in fields if str(field.get("name") or "").strip()}


def _field_candidates(profile: dict[str, Any], type_name: str) -> list[str]:
    fields = profile.get("fields") if isinstance(profile.get("fields"), list) else []
    return [
        str(field.get("name"))
        for field in fields
        if str(field.get("type") or "") == type_name and field.get("name")
    ][:8]


def workspace_data_overview(session_id: str) -> dict[str, Any]:
    root = workspace_dir(session_id)
    source_paths = [
        path for path in sorted(root.rglob("*"))
        if path.is_file() and not path.name.startswith(".") and "generated" not in path.relative_to(root).parts
    ]
    datasets: list[dict[str, Any]] = []
    field_sets: dict[str, set[str]] = {}
    for path in source_paths:
        rel = path.relative_to(root).as_posix()
        file_size = path.stat().st_size
        item: dict[str, Any] = {
            "name": path.name,
            "path": rel,
            "size": file_size,
            "category": classify_file(path),
        }
        if classify_file(path) == "table":
            if file_size <= 10 * 1024 * 1024 or path.suffix.lower() in {".csv", ".tsv", ".xlsx", ".xls", ".sqlite", ".db"}:
                try:
                    profile = profile_file(session_id, rel)
                    item["profile"] = _compact_profile(profile, field_limit=24)
                    fields = _field_names(profile)
                    field_sets[rel] = fields
                    item["date_fields"] = _field_candidates(profile, "日期")
                    item["numeric_fields"] = _field_candidates(profile, "数值")
                    item["categorical_fields"] = _field_candidates(profile, "分类")
                except Exception as exc:
                    item["profile_error"] = str(exc)
            else:
                item["profile_deferred"] = True
        datasets.append(item)
    relations: list[dict[str, Any]] = []
    paths = list(field_sets.keys())
    for left_index, left_path in enumerate(paths):
        for right_path in paths[left_index + 1:]:
            common = sorted(field_sets[left_path] & field_sets[right_path])
            if common:
                relations.append({"left": left_path, "right": right_path, "common_fields": common[:12]})
    return {
        "session_id": session_id,
        "datasets": datasets,
        "relations": relations,
        "generated_at": int(time.time()),
    }


def collect_file_info(session_id: str) -> str:
    items: list[str] = []
    root = workspace_dir(session_id)
    files = [
        path for path in sorted(root.rglob("*"))
        if path.is_file() and not path.name.startswith(".") and "generated" not in path.relative_to(root).parts
    ]
    for index, path in enumerate(files, start=1):
        size_kb = path.stat().st_size / 1024
        rel = path.relative_to(root).as_posix()
        info: dict[str, Any] = {"name": path.name, "path": rel, "size": f"{size_kb:.1f}KB"}
        ext = path.suffix.lower()
        peek: dict[str, Any] | None = None
        if is_delimited_text_file(path):
            preview = _preview_delimited_text(path, page=1, page_size=5)
            peek = {
                "columns": preview.get("columns", []),
                "sample_rows": _compact_rows(preview.get("rows", [])),
                "total_rows": preview.get("row_count"),
            }
        elif ext in {".csv", ".tsv"}:
            peek = _peek_csv(path)
        elif ext in {".xlsx", ".xls"}:
            peek = _peek_excel(path)
        elif ext in SQLITE_EXTENSIONS:
            peek = _peek_sqlite(path)
        if peek:
            info.update(peek)
        if classify_file(path) == "table":
            try:
                info["profile"] = _compact_profile(profile_file(session_id, rel), field_limit=16)
            except Exception:
                pass
        items.append(f"File {index}:\n{json.dumps(info, ensure_ascii=False, indent=2)}")
    return "\n\n".join(items)
